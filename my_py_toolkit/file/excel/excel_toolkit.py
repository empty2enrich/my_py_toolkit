# -*- coding: utf-8 -*-
# Copyright
# Author:
#
# cython: language_level=3
#

import openpyxl

from openpyxl import Workbook

def read_excel_values(file_path):
  """
  读取 excel 中的内容.
  Args:
    file_path:

  Returns:
    (dict): {sheet_name: val}
  """
  values = {}
  wb = openpyxl.load_workbook(file_path)
  for sheet_name in wb.sheetnames:
    sheet_val = []
    sheet = wb[sheet_name]
    col_num = sheet.max_column
    row_num = sheet.max_row
    for row_idx in range(1, row_num + 1):
      row_val = []
      for col_idx in range(1, col_num + 1):
        row_val.append(sheet.cell(row_idx, col_idx).value)
      sheet_val.append(row_val)
    values[sheet_name] = sheet_val
  return values

def write_excel(data, path, title="", sheet_index=0):
  """
  将表格数据写入 excel 里.
  Args:
    data:
    path:
    title:
    sheet_index:

  Returns:

  """
  wb = Workbook()
  sheet = wb.active if not title else wb.create_sheet(title, sheet_index)
  for row_index, row in enumerate(data):
    for col_index, cell in enumerate(row):
      sheet.cell(row=row_index + 1, column=col_index + 1).value = cell
  wb.save(path)
  wb.close()